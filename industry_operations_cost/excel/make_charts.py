# Industry Operations & Cost Optimization - Charts
# Author: Juan Jose Carrascal Pinzon
#
# Two charts straight from the workbook, so the analysis doesn't only live
# inside an .xlsx nobody opens. Reads operations_analysis.xlsx directly.
#
# Run: python make_charts.py

from pathlib import Path

import openpyxl
import matplotlib.pyplot as plt

THIS_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = THIS_DIR / "operations_analysis.xlsx"
IMAGES_DIR = THIS_DIR / "images"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def chart_oee_by_line(wb: openpyxl.Workbook) -> None:
    """OEE by Line sheet: header row 3, data from row 4.
    Columns: line_id, line_name, total_planned_min, total_downtime_min,
    total_units, total_defects, availability, quality, performance, oee"""
    ws = wb["OEE by Line"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0] is not None]
    lines = [r[1] for r in rows]
    oee = [r[9] * 100 for r in rows]

    fig, ax = plt.subplots(figsize=(7, 4.5))
    colors = ["#c0392b" if v < 80 else "#1f4e79" for v in oee]
    bars = ax.bar(lines, oee, color=colors, zorder=2)
    ax.bar_label(bars, fmt="%.1f%%", label_type="center", color="white", fontweight="bold", zorder=3)
    ax.axhline(85, color="#2e7d32", linestyle="--", linewidth=1.5, label="\"World class\" threshold (85%)", zorder=1)
    ax.set_title("OEE by Production Line")
    ax.set_ylabel("OEE (%)")
    ax.set_ylim(0, 100)
    ax.legend(loc="lower right")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(IMAGES_DIR / "oee_by_line.png", dpi=150)
    plt.close(fig)


def chart_cost_by_category(wb: openpyxl.Workbook) -> None:
    """Cost by Category sheet: header row 3, data from row 4.
    Columns: cost_type, total_cost"""
    ws = wb["Cost by Category"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0] is not None]
    rows.sort(key=lambda r: r[1], reverse=True)
    categories = [r[0] for r in rows]
    costs = [r[1] for r in rows]

    fig, ax = plt.subplots(figsize=(6.5, 4.5))
    bars = ax.bar(categories, costs, color="#1f4e79")
    ax.bar_label(bars, fmt="$%.0f", padding=3)
    ax.set_title("Total Cost by Category")
    ax.set_ylabel("Cost ($)")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(IMAGES_DIR / "cost_by_category.png", dpi=150)
    plt.close(fig)


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    print("Building OEE by line chart...")
    chart_oee_by_line(wb)

    print("Building cost by category chart...")
    chart_cost_by_category(wb)

    print(f"\nDone! Charts saved to {IMAGES_DIR}")


if __name__ == "__main__":
    main()
