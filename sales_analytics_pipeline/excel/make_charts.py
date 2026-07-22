# Sales Analytics Pipeline - Charts
# Author: Juan Jose Carrascal Pinzon
#
# Two charts straight from the workbook, so the analysis doesn't only live
# inside an .xlsx nobody opens. Reads sales_analysis.xlsx directly.
#
# Run: python make_charts.py

from pathlib import Path

import openpyxl
import matplotlib.pyplot as plt

THIS_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = THIS_DIR / "sales_analysis.xlsx"
IMAGES_DIR = THIS_DIR / "images"
IMAGES_DIR.mkdir(parents=True, exist_ok=True)


def chart_monthly_revenue(wb: openpyxl.Workbook) -> None:
    """Monthly Revenue sheet: header row 3, data from row 4.
    Columns: month, total_revenue, total_orders, prev_month_revenue, growth_pct"""
    ws = wb["Monthly Revenue"]
    rows = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0] is not None]
    months = [r[0] for r in rows]
    revenue = [r[1] for r in rows]

    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.plot(months, revenue, marker="o", color="#1f4e79")
    ax.set_title("Monthly Revenue")
    ax.set_ylabel("Revenue ($)")
    ax.set_xlabel("Month")
    ax.tick_params(axis="x", rotation=60)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(IMAGES_DIR / "monthly_revenue.png", dpi=150)
    plt.close(fig)


def chart_top_customers(wb: openpyxl.Workbook) -> None:
    """Top Customers sheet: header row 4, data from row 5.
    Columns: rank, customer_name, segment, total_orders, total_revenue"""
    ws = wb["Top Customers"]
    rows = [r for r in ws.iter_rows(min_row=5, max_row=9, values_only=True) if r[0] is not None]
    labels = [r[1] for r in rows]
    revenue = [r[4] for r in rows]

    fig, ax = plt.subplots(figsize=(7, 4.5))
    bars = ax.barh(labels[::-1], revenue[::-1], color="#1f4e79")
    ax.bar_label(bars, fmt="$%.0f", padding=3)
    ax.set_title("Top 5 Customers by Revenue")
    ax.set_xlabel("Revenue ($)")
    fig.tight_layout()
    fig.savefig(IMAGES_DIR / "top_customers.png", dpi=150)
    plt.close(fig)


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

    print("Building monthly revenue chart...")
    chart_monthly_revenue(wb)

    print("Building top customers chart...")
    chart_top_customers(wb)

    print(f"\nDone! Charts saved to {IMAGES_DIR}")


if __name__ == "__main__":
    main()
